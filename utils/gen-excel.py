#!/usr/bin/env python3

import os
import click
import pathlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# List of CSV filenames in the desired order
csv_files = [
    'Abbreviations.csv',
    'AreaCodeValues.csv',
    'CommodityValues.csv',
    'ConstructionValues.csv',
    'CountryValues.csv',
    'CoverageValues.csv',
    'CurrencyValues.csv',
    'FinancialCodeValues.csv',
    'IndustryCodeValues.csv',
    'OccupancyValues.csv',
    'OEDCRFieldAppendix.csv',
    'OEDInputFields.csv',
    'OtherValues.csv',
    'PerilsCovered.csv',
    'PerilValues.csv',
    'Versioning.csv',
]


# Directory containing the CSV files
source_csv_default = pathlib.Path(os.path.dirname(os.path.realpath(__file__))).parent.joinpath('OpenExposureData')



@click.command()
@click.option( "--source-csv-dir", required=True, default=source_csv_default, help="Path to dir with OED csv files")
@click.option( "--output-path", default="OpenExposureData.xlsx", help="Path to write json oed file")
def oed_spec_to_excel(output_path, source_csv_dir):

    # Create a Pandas Excel writer using openpyxl as the engine
    pathlib.Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    combined_workbook_path = pathlib.Path(output_path).absolute()
    excel_writer = pd.ExcelWriter(combined_workbook_path, engine='openpyxl')

    # Iterate over the list of CSV filenames
    for filename in csv_files:
        # Construct the full file path
        file_path = os.path.join(source_csv_dir, filename)

        # Check if the file exists
        if os.path.exists(file_path):
            # Read the CSV file into a DataFrame
            df = pd.read_csv(file_path,dtype=str, keep_default_na=False, na_values=[])

            # Use the filename (without extension) as the sheet name
            sheet_name = os.path.splitext(filename)[0]

            # Write the DataFrame to a sheet in the Excel workbook
            df.to_excel(excel_writer, sheet_name=sheet_name, index=False, na_rep='NaN', header=True)
        else:
            print(f"File {file_path} does not exist.")

    # Save the workbook
    excel_writer.close()

    # Load the workbook with openpyxl to add table formatting
    wb = load_workbook(combined_workbook_path)

    # Iterate over the list of CSV filenames again to add table formatting
    for filename in csv_files:
        sheet_name = os.path.splitext(filename)[0]

        # Get the sheet from the workbook
        ws = wb[sheet_name]

        # Define the table range (adjust if needed)
        table_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

        # Create an Excel table
        tab = Table(displayName=sheet_name, ref=table_range)

        # Add a default style with striped rows
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Save the workbook again to apply table formatting
    wb.save(combined_workbook_path)
    print(f"Excel Workbook has been created with each CSV in a separate sheet formatted as a table. Stored in '{combined_workbook_path}'")



if __name__ == "__main__":
    oed_spec_to_excel()
